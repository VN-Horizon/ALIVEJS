import openpyxl
import struct

def load_columns_from_xlsx(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the active worksheet
    sheet = workbook.active
    second_column = []
    third_column = []

    # Iterate through the rows and extract data from the second and third columns
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=4):
        try:
            r2 = row[1].value
            second_column.append(r2)  # Second column
        except:
            second_column.append("")
        try:
            r3 = row[2].value
            third_column.append(r3)  # Second column
        except:
            third_column.append("")

    return second_column, third_column

# Example usage
if __name__ == "__main__":
    file_path = "机翻初版.xlsx"  # Replace with your .xlsx file path
    second_col, third_col = load_columns_from_xlsx(file_path)
    indexes_to_delete = []
    for i in range(len(second_col)):
        if not second_col[i] or not third_col[i]:
            indexes_to_delete.append(i)

    for i in reversed(indexes_to_delete):
        del second_col[i]
        del third_col[i]

    with open("translation.bin", 'wb') as f:
        for i in range(len(second_col)):
            # Encode strings as UTF-8 and write their lengths followed by the encoded data
            second_encoded = second_col[i].encode('utf-8')
            third_encoded = third_col[i].encode('utf-8')

            # Write the length of each string followed by the string itself
            f.write(struct.pack('I', len(second_encoded)))  # Length of second column string
            f.write(second_encoded)  # Encoded second column string
            f.write(struct.pack('I', len(third_encoded)))  # Length of third column string
            f.write(third_encoded)  # Encoded third column string
